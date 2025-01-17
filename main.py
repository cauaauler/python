from openpyxl import load_workbook
from openpyxl import Workbook

# Carregar o arquivo Excel enviado
file_path = './Excel_100_Rows.xlsx'
workbook = load_workbook(file_path)
sheet = workbook.active

# Ler os dados da planilha e organizar em blocos baseados no valor "name"
data_blocks = []
current_block = []

for row in sheet.iter_rows(values_only=True):
    if "Name" in row:
        if current_block:  # Salvar bloco anterior antes de iniciar um novo
            data_blocks.append(current_block)
        current_block = []  # Iniciar novo bloco
    if row and any(cell is not None and cell != '' for cell in row):  # Verifica se a linha não está vazia
        current_block.append([row[0], row[2]])


# Adicionar o último bloco, se existir
if current_block:
    data_blocks.append(current_block)

# Criar um novo arquivo XLSX com os dados organizados
new_workbook = Workbook()
new_sheet = new_workbook.active
new_sheet.title = "Processed Data"

# Escrever os dados no novo arquivo
row_index = 1
for block in data_blocks:
    for row in block:
        for col_index, cell in enumerate(row, start=1):
            new_sheet.cell(row=row_index, column=col_index, value=str(cell) if cell is not None else "")
        row_index += 1

    row_index += 1  # Linha extra entre blocos

# Salvar o arquivo
xlsx_file_path = './processed_data.xlsx'
new_workbook.save(xlsx_file_path)
