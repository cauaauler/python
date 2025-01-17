from openpyxl import load_workbook
from openpyxl import Workbook

# Carregar o arquivo Excel enviado
file_path = r'./Financial Sample.xlsx'
workbook = load_workbook(file_path)
sheet = workbook.active

# Ler os dados da planilha e organizar em blocos baseados no valor "name"
data_blocks = []
current_block = []

for row in sheet.iter_rows(values_only=True):
    if "Name" in row:
        if current_block:  # Salvar bloco anterior antes de iniciar um novo
            data_blocks.append(current_block)
        current_block = []  # Iniciar novo bloco
    if row and any(cell is not None and cell != '' for cell in row):  # Verifica se a linha não está vazia
        current_block.append([row[1], row[9]])


# Adicionar o último bloco, se existir
if current_block:
    data_blocks.append(current_block)

# Criar um novo arquivo XLSX com os dados organizados
new_workbook = Workbook()
new_sheet = new_workbook.active
new_sheet.title = "Processed Data"

# Escrever os dados no novo arquivo
row_index = 1
for block in data_blocks:
    for row in block:
        for col_index, cell in enumerate(row, start=1):
            new_sheet.cell(row=row_index, column=col_index, value=str(cell) if cell is not None else "")
        row_index += 1

    row_index += 1  # Linha extra entre blocos

# Salvar o arquivo
xlsx_file_path = r'./processed_data.xlsx'
new_workbook.save(xlsx_file_path)
from openpyxl import load_workbook, Workbook

# Carregar o arquivo Excel processado
file_path = './processed_data.xlsx'
workbook = load_workbook(file_path)
sheet = workbook.active

# Lista para armazenar os países e valores
countries = []

# Processar as linhas do arquivo
for row in sheet.iter_rows(min_row=2, values_only=True) :
    if row and row[0] and row[1]:  # Verifica se a linha não está vazia e tem os dois valores necessários
        country_name = row[0]
        country_value = float(row[1])
        
        # Verificar se o país já está na lista
        found = False
        for country in countries:
            if country['name'] == country_name:
                country['value'] += country_value
                found = True
                break
        
        # Adicionar novo país, se ainda não estiver na lista
        if not found:
            countries.append({'name': country_name, 'value': country_value})

# Criar um novo arquivo XLSX com os dados organizados
new_workbook = Workbook()
new_sheet = new_workbook.active
new_sheet.title = "Processed Data"

# Escrever os dados no novo arquivo
row_index = 1
for country in countries:
    new_sheet.cell(row=row_index, column=1, value=country['name'])
    new_sheet.cell(row=row_index, column=2, value=country['value'])
    row_index += 1

# Salvar o arquivo
xlsx_file_path = './final_processed_data.xlsx'
new_workbook.save(xlsx_file_path)

print(f"Arquivo processado e salvo em: {xlsx_file_path}")
